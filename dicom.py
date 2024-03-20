
import os
import shutil
import pydicom
import time

import numpy as np
import matplotlib.pyplot as plt

import openpyxl

def read_excel(file_path, columns):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # This gets the first sheet by default

    # Alternatively, you can specify the sheet by name
    # sheet = workbook['Sheet1']  # Replace 'Sheet1' with the name of your sheet

    # List to store values from specified columns in each row
    values_list = []

    # Iterate through rows
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        # Access cell values in specified columns
        cell_values = [row[column - 1] for column in columns]  # Adjust column index to 0-based
        print(cell_values)
        # time.sleep(10)
        # Append values from specified columns to the list
        values_list.append(cell_values)

    workbook.close()

    return values_list

def mark_point_on_slice(dcm_file, point_mm):
    """
    Mark a point on a DICOM slice.
    """
    # Read the DICOM file
    ds = pydicom.dcmread(dcm_file)
    
    # Extract relevant metadata
    image_position = ds.ImagePositionPatient
    image_orientation = ds.ImageOrientationPatient
    pixel_spacing = ds.PixelSpacing
    
    # Convert physical coordinates (mm) to pixel coordinates
    point_px = np.round(np.divide(np.subtract(point_mm, image_position[:2]), pixel_spacing)).astype(int)
    
    # Read the pixel data
    pixel_data = ds.pixel_array
    
    # Plot the DICOM image
    plt.imshow(pixel_data, cmap='gray')
    
    # Mark the point on the image
    plt.plot(point_px[0], point_px[1], 'ro', markersize=5)  # Assuming point_px is (x, y)
    
    # Show the image with the marked point
    plt.show()

def copy_and_rename_folder(src_folder):
    """
    Copy a folder and add "exported_" to the new folder name.
    Returns the new folder path.
    """
    # Extract the folder name from the source folder path
    folder_name = os.path.basename(src_folder)
    
    # Add "exported_" to the folder name
    new_folder_name = "exported_" + folder_name
    
    # Get the path to the source folder's parent directory
    parent_dir = os.path.dirname(src_folder)
    
    # Create the new destination folder path with the new name
    new_dest_folder = os.path.join(parent_dir, new_folder_name)
    
    # Remove the existing folder if it exists
    if os.path.exists(new_dest_folder):
        shutil.rmtree(new_dest_folder)
    
    # Copy the folder to the destination with the new name
    shutil.copytree(src_folder, new_dest_folder)
    
    # Replace backslashes with forward slashes
    new_dest_folder = new_dest_folder.replace("\\", "/")
    
    return new_dest_folder


def get_dcm_file_paths(directory):
    """
    Get all file paths with the extension ".dcm" in the specified directory and its subdirectories.
    Returns a tuple containing the count of files and the list of file paths.
    """
    dcm_file_paths = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith(".dcm"):
                dcm_file_paths.append(os.path.join(root, file))
    return len(dcm_file_paths), dcm_file_paths
def calculate_pixel_positions(dcm_file):
    """
    Calculate the position of each pixel in a DICOM image.
    """
    # Read the DICOM file
    ds = pydicom.dcmread(dcm_file)
    
    # Extract relevant metadata attributes
    image_position = np.array(ds.ImagePositionPatient)
    image_orientation = np.array(ds.ImageOrientationPatient)
    pixel_spacing = np.array(ds.PixelSpacing)
    rows, columns = ds.Rows, ds.Columns
    
    # Compute the normal vector to the image plane
    slice_normal = np.cross(image_orientation[:3], image_orientation[3:])
    
    # Initialize arrays to store pixel positions
    pixel_positions = np.zeros((rows, columns, 3))
    
    # Calculate pixel positions
    for i in range(rows):
        for j in range(columns):
            # Compute pixel position relative to the image position
            pixel_position_relative = (i * pixel_spacing[0] * image_orientation[:3] +
                                       j * pixel_spacing[1] * image_orientation[3:])
            # Adjust pixel position based on image position
            pixel_positions[i, j] = image_position + pixel_position_relative
    
    return pixel_positions

def find_pixel_from_position(dcm_file, target_position, tolerance=1e-2):
    """
    Find the pixel corresponding to a given position in a DICOM image.
    """
    # Read the DICOM file
    ds = pydicom.dcmread(dcm_file)
    
    # Extract relevant metadata attributes
    image_position = np.array(ds.ImagePositionPatient)
    image_orientation = np.array(ds.ImageOrientationPatient)
    pixel_spacing = np.array(ds.PixelSpacing)
    rows, columns = ds.Rows, ds.Columns
    
    # Initialize arrays to store pixel positions
    pixel_positions = np.zeros((rows, columns, 3))
    
    # Calculate pixel positions
    for i in range(rows):
        for j in range(columns):
            # Compute pixel position relative to the image position
            pixel_position_relative = (i * pixel_spacing[0] * image_orientation[:3] +
                                       j * pixel_spacing[1] * image_orientation[3:])
            # Adjust pixel position based on image position
            pixel_positions[i, j] = image_position + pixel_position_relative
    
    # Find the pixel closest to the target position
    min_distance = float('inf')
    closest_pixel = None
    for i in range(rows):
        for j in range(columns):
            distance = np.linalg.norm(pixel_positions[i, j] - target_position)
            if distance < min_distance:
                min_distance = distance
                closest_pixel = (i, j)
    
    # Check if the closest pixel is within the tolerance
    if min_distance <= tolerance:
        return closest_pixel
    else:
        return None

def modify_pixel_data_for_centers(dcm_file, center1, center2, region_size):
    """
    Modify the pixel data in a DICOM file around specified centers.

    Args:
        dcm_file (str): Path to the DICOM file.
        center1 (tuple): Tuple containing the (x, y) coordinates of the first center of the region.
        center2 (tuple): Tuple containing the (x, y) coordinates of the second center of the region.
        region_size (int): Size of the region (square) around the centers.
    """
    # Read the DICOM file
    ds = pydicom.dcmread(dcm_file)
    arr = ds.pixel_array
    arr_origin = ds.pixel_array
    m, n = arr.shape

    # Modify pixel data around center1
    if center1 is not None:

        x1, y1 = center1
        for i in range(m):
            for j in range(n):
                if (((i >= x1 - region_size) and (i < x1 + region_size)) and 
                    ((j >= y1 - region_size) and (j < y1 + region_size))):
                    arr[i][j] = 0 

    # Modify pixel data around center2
    if center2 is not None:
        x2, y2 = center2
        for i in range(m):
            for j in range(n):
                if (((i >= x2 - region_size) and (i < x2 + region_size)) and 
                    ((j >= y2 - region_size) and (j < y2 + region_size))):
                    arr[i][j] = 0 

    # Save the modified pixel data back to the DICOM dataset
    # Check if arrays are equal
    if np.array_equal(arr, arr_origin):
        print("Arrays are the same.")
        time.sleep(20)
    else:
        print("Arrays are different.")
        time.sleep(20)
    if arr.all() != arr_origin.all():
        print("sdfsssssssssss")
        time.sleep(20)
        ds.PixelData = arr.tobytes()
        # Save the DICOM file with the modified pixel data
        modified_file_path = dcm_file.replace('.dcm', '_modified.dcm')
        ds.save_as(modified_file_path)
        return True
    print("aaaaaaaaaaaaaaaaaaaaaa")
    time.sleep(20)
    return False

xlsx_path = 'Only_Targeted_Biopsies.xlsx'
columns_to_read = [9, 10, 11, 12, 13, 14, 24, 25]  # Example: Read values from columns A, C, and E
MRI_values = read_excel(xlsx_path, columns_to_read)
print(MRI_values)
time.sleep(5)


src_folder = "D:/DevProject/3DSlicer/dev/DICOM_MRML/Prostate-MRI-US-Biopsy"
new_folder_path = copy_and_rename_folder(src_folder)
print("New folder path:", new_folder_path)
time.sleep(5)

region_size = 20  # Example region size

for MRI_value in MRI_values:
    num_files, dcm_files = get_dcm_file_paths(new_folder_path)
    print("Number of .dcm files:", num_files)
    for index, file_path in enumerate(dcm_files, start=1):
        print(f"{index}. {file_path}")
        ds = pydicom.dcmread(file_path)
        print(ds.Modality)
        if(ds.Modality == "MR"):
            print(ds.PatientID)
            print(ds.SeriesInstanceUID)
            # print(ds.Modality)
            # print(ds.InstanceNumber)
            print(ds.ImagePositionPatient)
            print(ds.ImageOrientationPatient)
            print(ds.PixelSpacing)

            # pixel_positions = calculate_pixel_positions(file_path)
            # print("Pixel Positions:", pixel_positions)
            # # Assuming pixel_positions is a 2D array containing coordinates
            # for row in pixel_positions:
            #     for pixel in row:
            #         print("Coordinate:", pixel)

            # time.sleep(20)

            # if((ds.PatientID == MRI_value[6]) and (ds.SeriesInstanceUID == MRI_value[7])):
                # Example usage:
                # target_position1 = np.array([MRI_value[0], MRI_value[1], MRI_value[2]])
                # target_position2 = np.array([MRI_value[3], MRI_value[4], MRI_value[5])
            target_position1 = np.array([-112.30713844,   69.330513,    -64.83107376])
            target_position2 = np.array([-120.43213844,   33.580513,    -64.83107376])
            pixel1 = find_pixel_from_position(file_path, target_position1)
            pixel2 = find_pixel_from_position(file_path, target_position2)
            print((ds.pixel_array).shape)
            print("Pixel Corresponding to Target Position:", pixel1, pixel2)
            modify_pixel_data_for_centers(file_path, pixel1, pixel2, region_size)


 

